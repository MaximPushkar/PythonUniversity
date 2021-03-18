import cgi
import openpyxl
import random
from string import Template

RESULT_LINE = """
      <p>$TeamName</p>
"""

RESULT_ROW = """
        <tr>
            <th>$a</th>
            <th>$b</th>
            <th>$c</th>
            <th>$d</th>
            <th>$e</th>
            <th>$f</th>
            <th>$g</th>
            <th>$h</th>
            <th>$i</th>
        </tr>
"""

NAME_OPTION = """
<option value="$TeamName">$TeamName</option>
"""

STATUS = """
<h1>$status</h1>
"""


class MatchResults:

    def __init__(self, data1, data2):
        self._names = data1
        self._match = data2

    def get_names(self):
        wb = openpyxl.load_workbook(self._names)
        ws = wb.worksheets[0]
        return [row[0].value for row in ws.rows]

    def get_matches(self):
        wb = openpyxl.load_workbook(self._match)
        ws = wb.worksheets[0]
        return [[int(cell.value) for cell in row] for row in ws.rows]

    def get_data(self):
        names = self.get_names()
        data = dict(zip(list(range(1, len(names) + 1)), [[name, 0, 0, 0, 0, 0, 0, 0] for name in names]))
        # (номер : [ім'я, ігор, виграшів, нічиїх, поразок, м’ячів забито, м’ячів пропущено, очок]

        for match in self.get_matches():
            player_1, player_2, balls_1, balls_2 = match
            data[player_1][1] += 1
            data[player_2][1] += 1

            data[player_1][5] += balls_1
            data[player_2][5] += balls_2
            data[player_1][6] += balls_2
            data[player_2][6] += balls_1

            if balls_1 > balls_2:
                data[player_1][2] += 1
                data[player_2][4] += 1

                data[player_1][7] += 3

            elif balls_1 < balls_2:
                data[player_1][4] += 1
                data[player_2][2] += 1

                data[player_2][7] += 3

            else:
                data[player_1][3] += 1
                data[player_2][3] += 1

                data[player_1][7] += 1
                data[player_2][7] += 1

        # return data
        order = sorted(data, key=lambda team: (data[team][7], data[team][5] - data[team][6],
                                               data[team][5], random.random()), reverse=True)
        return [data[i] for i in order]

    def __call__(self, environ, start_response):
        path = environ.get("PATH_INFO", "").lstrip("/")
        params = {"lines": "", "rows": "", "names": "", "status": ""}
        status = "200 OK"
        headers = [("Content-Type", "text/html; charset=utf-8")]

        # http://127.0.0.1:8000/
        if path == "":
            names = ""
            for team in self.get_data():
                names += Template(NAME_OPTION).substitute(TeamName=str(team[0]))
            params["names"] = names
            html = "templates/add_result.html"

        # http://127.0.0.1:8000/operation_status
        elif path == "operation_status":
            form = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ)
            name_1 = form.getfirst("from", "")
            name_2 = form.getfirst("to", "")
            balls_1 = form.getfirst("balls_1", "")
            balls_2 = form.getfirst("balls_2", "")

            if name_1 and name_2 and balls_1 and balls_2:
                try:
                    names = self.get_names()
                    pos_1 = names.index(name_1)
                    pos_2 = names.index(name_2)
                    wb = openpyxl.load_workbook(self._match)
                    ws = wb.worksheets[0]
                    ws.append([pos_1 + 1, pos_2 + 1, int(balls_1), int(balls_2)])
                    wb.save(self._match)
                    response = "Ваш результат додано до файлу з результатами"
                except Exception as e:
                    response = "Виникла помилка " + str(e)
                    print(e)
                params["status"] = Template(STATUS).substitute(status=response)

            # Якщо у формі задано недостатньо параметрів, перенаправляємо на головну сторінку
            else:
                status = "303 SEE OTHER"
                headers.append(("Location", "/"))
            html = "templates/operation_status.html"


        # http://127.0.0.1:8000/partial
        elif path == "partial":
            lines = ""
            counter = 0
            for team in self.get_data():
                counter += 1
                lines += Template(RESULT_LINE).substitute(TeamName=str(counter) + ")  " + str(team[0]))
            params["lines"] = lines
            html = "templates/just_results.html"

        # http://127.0.0.1:8000/full
        elif path == "full":
            rows = ""
            counter = 0
            for team in self.get_data():
                counter += 1
                rows += Template(RESULT_ROW).substitute(a=str(counter), b=team[0], c=team[1],
                                                        d=team[2], e=team[3], f=team[4],
                                                        g=team[5], h=team[6], i=team[7])
            params["rows"] = rows
            html = "templates/full_results.html"

        # http://127.0.0.1:8000/<будь-який інший запит>
        else:
            status = "404 NOT FOUND"
            html = "templates/error_404.html"

        start_response(status, headers)
        with open(html, encoding="utf-8") as f:
            page = Template(f.read()).substitute(params)
        return [bytes(page, encoding="utf-8")]


HOST = ""
PORT = 8000

if __name__ == '__main__':
    app = MatchResults("Data/names.xlsx", "Data/results.xlsx")
    from wsgiref.simple_server import make_server

    print(" === Local webserver === ")
    make_server(HOST, PORT, app).serve_forever()
