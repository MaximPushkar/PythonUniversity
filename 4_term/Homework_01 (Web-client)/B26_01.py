# Скласти програму, яка читає прогноз погоди у заданому місті з сайту
# sinoptik.ua та зберігає у файлі Excel у окремому рядку поточну дату та прогнози
# максимальної та мінімальної температури на кожний з наступних 5 днів.
# Запит на погоду у заданому місті:
# https://sinoptik.ua/погода-<місто>/
# Наприклад, https://sinoptik.ua/погода-киев


from urllib.request import urlopen  # Функція для отримання веб-сторінки з мережі
from urllib.parse import quote  # Функція для кодування кирилиці в url
import re
import openpyxl

DATE = r'<p class="date .{0,9}">(?P<DATE_NAME>.+)</p>'
MONTH = r'<p class="month">(?P<MONTH_NAME>.+)</p>'
TEMP = r'<span>(?P<DEG>.+)&deg;</span>'


def find_temperature(html):
    re_list = re.findall(TEMP, html)
    return re_list


def find_date(html):
    re_list = re.findall(DATE, html)
    return re_list


def find_month(html):
    re_list = re.findall(MONTH, html)
    return re_list


def make_url(city):
    main_url = 'https://sinoptik.ua/'
    path = 'погода-' + city
    return main_url + quote(path, encoding="utf-8")


def get_html(url):
    """ Повертає розкодавані дані веб-сторінки за заданою адресою."""
    return str(urlopen(url).read(), encoding="utf-8", errors="ignore")


def fix_this_awful_html(html):
    """ Весь код написано в один рядок (!) Шукаємо кінці рядків вигляду '...> '   """
    html = '> \n'.join(html.split('> '))
    return html


def main_function(city, file):
    url = make_url(city)
    html = get_html(url)
    html = fix_this_awful_html(html)
    # print(html)

    temp_list, date_list, month_list = find_temperature(html), find_date(html), find_month(html)
    # print(temp_list, date_list, month_list)

    wb = openpyxl.Workbook()
    # wb.create_sheet(title="sheet", index=0)
    # print(wb.sheetnames)
    ws = wb["Sheet"]

    ws.append(["Дата", "Мінімальна температура", "Максимальна температура"])

    for i in range(1, 6):
        ws.append([date_list[i - 1] + ' ' + month_list[i - 1], temp_list[2 * i - 2], temp_list[2 * i - 1]])

    wb.save("output.xlsx")


if __name__ == '__main__':
    main_function('киев', 'output.xlsx')
    print()
